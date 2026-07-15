"""
export_xlsx.py

Purpose:
    Builds the XLSX spreadsheet required by your assignment (Step 4c),
    with exactly these columns:
        repository_id, project_type, project_title, primary_class,
        secondary_class, no_project_files

    One row per project (all 22 - not just the 18 QD_PROJECTs, since
    NOT_A_PROJECT rows still need to appear with empty primary/secondary
    class columns).

How to use it:
    python export_xlsx.py

    Produces: classification_results.xlsx (in the classification folder)
"""

import sqlite3
from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill
from openpyxl.utils import get_column_letter

BASE_DIR = Path(__file__).resolve().parent.parent
DB_PATH = BASE_DIR / "23158572-sq26-classification.db"
OUTPUT_PATH = Path(__file__).resolve().parent / "classification_results.xlsx"

COLUMNS = [
    "repository_id",
    "project_type",
    "project_title",
    "primary_class",
    "secondary_class",
    "no_project_files",
]


def simplify_class_label(full_label):
    """
    Classifications are stored like:
      'A01 - Crop and animal production... (Section A: Agriculture...)'
    For the spreadsheet, drop the trailing '(Section ...)' part to keep
    cells short and readable - the code + title is enough.
    """
    if not full_label:
        return ""
    if " (Section " in full_label:
        return full_label.split(" (Section ")[0]
    return full_label


def get_project_rows(conn):
    """
    One row per project, joining in:
      - its PROJECT-level classification (if any)
      - its total file count (all files, any status)
    """
    query = """
        SELECT
            p.id,
            p.repository_id,
            p.type AS project_type,
            p.title AS project_title,
            c.primary_class,
            c.secondary_class,
            (SELECT COUNT(*) FROM files f WHERE f.project_id = p.id)
                AS no_project_files
        FROM projects p
        LEFT JOIN classifications c
            ON c.project_id = p.id AND c.target_type = 'PROJECT'
        ORDER BY p.id
    """
    return conn.execute(query).fetchall()


def build_workbook(rows):
    wb = Workbook()
    ws = wb.active
    ws.title = "Classification Results"

    # Header row
    header_font = Font(name="Arial", bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="4472C4", end_color="4472C4",
                               fill_type="solid")
    for col_idx, col_name in enumerate(COLUMNS, start=1):
        cell = ws.cell(row=1, column=col_idx, value=col_name)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center")

    # Data rows
    body_font = Font(name="Arial")
    for row_idx, row in enumerate(rows, start=2):
        (project_id, repository_id, project_type, project_title,
         primary_class, secondary_class, no_project_files) = row

        values = [
            repository_id,
            project_type,
            project_title,
            simplify_class_label(primary_class),
            simplify_class_label(secondary_class),
            no_project_files,
        ]
        for col_idx, value in enumerate(values, start=1):
            cell = ws.cell(row=row_idx, column=col_idx, value=value)
            cell.font = body_font

    # Reasonable column widths so text isn't cut off
    widths = [14, 16, 45, 55, 55, 16]
    for col_idx, width in enumerate(widths, start=1):
        ws.column_dimensions[get_column_letter(col_idx)].width = width

    ws.freeze_panes = "A2"  # keep header visible while scrolling
    return wb


def run():
    conn = sqlite3.connect(DB_PATH)
    rows = get_project_rows(conn)
    conn.close()

    if not rows:
        print("No projects found - did you run the earlier scripts?")
        return

    wb = build_workbook(rows)
    wb.save(OUTPUT_PATH)

    print(f"Saved {len(rows)} project rows to: {OUTPUT_PATH}")
    print("\nColumns: " + ", ".join(COLUMNS))
    print("\nFirst 3 rows preview:")
    for row in rows[:3]:
        print(f"  {row}")


if __name__ == "__main__":
    run()